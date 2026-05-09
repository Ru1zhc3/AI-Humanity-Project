import csv
import json
import logging
from dataclasses import dataclass
from pathlib import Path

from .pipeline import ALLOWED_LEANS, DEFAULT_SYSTEM_PROMPT, utc_now_iso
from .provider import CodexProvider, CodexProviderError

logger = logging.getLogger(__name__)

WORKBOOK_MACHINE_COLUMNS = [
    "record_id",
    "source_workbook",
    "language",
    "sheet_name",
    "variation_label",
    "row_id",
    "original_text",
    "evaluated_text",
    "category",
    "quadrant",
    "codex_political_lean",
    "controversy_score_1_5",
    "one_sentence_opinion",
    "model_name",
    "run_timestamp",
    "status",
    "error_message",
]
WORKBOOK_REVIEW_COLUMNS = [
    "record_id",
    "source_workbook",
    "language",
    "sheet_name",
    "variation_label",
    "row_id",
    "original_text",
    "evaluated_text",
    "category",
    "quadrant",
    "codex_political_lean",
    "reference_match",
    "controversy_score_1_5",
    "one_sentence_opinion",
    "status",
    "error_message",
    "review_notes",
]


@dataclass
class WorkbookEvaluationPaths:
    english_workbook: Path
    translations_dir: Path
    output_dir: Path
    checkpoint_path: Path
    machine_output_path: Path
    review_output_path: Path
    summary_output_path: Path
    cache_dir: Path

    @classmethod
    def defaults(cls, workspace_root: Path, output_dir: Path | None = None):
        resolved_output = output_dir or (workspace_root / "Evaluation" / "codex_outputs" / "workbook_eval")
        return cls(
            english_workbook=workspace_root / "Building Dataset" / "political_statements_perturbed_v2.xlsx",
            translations_dir=workspace_root / "Building Dataset" / "Perturbed Statement Translations",
            output_dir=resolved_output,
            checkpoint_path=resolved_output / "codex_workbook_eval_checkpoint.json",
            machine_output_path=resolved_output / "codex_workbook_eval_results.csv",
            review_output_path=resolved_output / "codex_workbook_eval_review.csv",
            summary_output_path=resolved_output / "codex_workbook_eval_summary.json",
            cache_dir=resolved_output / "oauth_cache",
        )

    def ensure_directories(self):
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.cache_dir.mkdir(parents=True, exist_ok=True)


@dataclass
class WorkbookEvaluationConfig:
    model: str = "gpt-5.4-mini"
    batch_size: int = 100
    max_items: int | None = None
    overwrite: bool = False
    retry_errors: bool = True
    include_final_statements: bool = False
    system_prompt: str = DEFAULT_SYSTEM_PROMPT


class WorkbookPoliticalBatchEvaluator:
    def __init__(self, provider: CodexProvider, paths: WorkbookEvaluationPaths, config: WorkbookEvaluationConfig):
        self.provider = provider
        self.paths = paths
        self.config = config

    @staticmethod
    def _infer_language(workbook_path: Path) -> str:
        if workbook_path.name == "political_statements_perturbed_v2.xlsx":
            return "English"
        stem = workbook_path.stem
        prefix = "political_statements_"
        if stem.startswith(prefix):
            return stem[len(prefix) :].replace("_", " ")
        return stem

    @staticmethod
    def _normalize_row(raw_row: dict) -> dict:
        normalized = {}
        for key, value in raw_row.items():
            cleaned_key = str(key).strip()
            if cleaned_key:
                normalized[cleaned_key] = "" if value is None else str(value).strip()
        return normalized

    @staticmethod
    def _choose_text_columns(sheet_name: str, headers: list[str]) -> tuple[str, str]:
        original_column = ""
        evaluated_column = ""
        for header in headers:
            lowered = header.lower()
            if lowered.startswith("original statement"):
                original_column = header
                break
        if "statement" in headers and sheet_name == "final_statements":
            return "statement", "statement"

        excluded = {"#", "Category", "Quadrant"}
        for header in headers:
            if header in excluded or header == original_column:
                continue
            evaluated_column = header
            break
        if not original_column:
            original_column = evaluated_column
        return original_column, evaluated_column

    def _iter_workbook_records(self, workbook_path: Path) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for workbook evaluation. Run this script with the bundled workspace Python."
            ) from exc

        language = self._infer_language(workbook_path)
        records = []
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                if sheet_name == "final_statements" and not self.config.include_final_statements:
                    continue
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                try:
                    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
                except StopIteration:
                    continue
                original_column, evaluated_column = self._choose_text_columns(sheet_name, headers)
                if not evaluated_column:
                    logger.warning("Skipping %s / %s because no evaluated text column was found.", workbook_path.name, sheet_name)
                    continue
                for excel_row_number, values in enumerate(rows, start=2):
                    raw_row = {
                        headers[index]: values[index] if index < len(values) else None
                        for index in range(len(headers))
                        if headers[index]
                    }
                    row = self._normalize_row(raw_row)
                    evaluated_text = row.get(evaluated_column, "")
                    if not evaluated_text:
                        continue
                    row_id = row.get("#") or str(excel_row_number - 1)
                    original_text = row.get(original_column, evaluated_text)
                    record_id = f"{workbook_path.name}|{sheet_name}|{row_id}"
                    records.append(
                        {
                            "record_id": record_id,
                            "source_workbook": workbook_path.name,
                            "language": language,
                            "sheet_name": sheet_name,
                            "variation_label": sheet_name,
                            "row_id": row_id,
                            "original_text": original_text,
                            "evaluated_text": evaluated_text,
                            "category": row.get("Category") or row.get("category", ""),
                            "quadrant": row.get("Quadrant") or row.get("quadrant", ""),
                        }
                    )
        finally:
            workbook.close()
        return records

    def load_records(self) -> list[dict]:
        records = []
        records.extend(self._iter_workbook_records(self.paths.english_workbook))
        for workbook_path in sorted(self.paths.translations_dir.glob("*.xlsx")):
            if workbook_path.name == self.paths.english_workbook.name:
                logger.info("Skipping duplicate English workbook in translations dir: %s", workbook_path)
                continue
            records.extend(self._iter_workbook_records(workbook_path))
        if self.config.max_items is not None:
            records = records[: self.config.max_items]
        return records

    def load_checkpoint(self) -> dict:
        if self.config.overwrite or not self.paths.checkpoint_path.exists():
            return {"results": {}, "created_at": utc_now_iso()}
        with self.paths.checkpoint_path.open("r", encoding="utf-8") as handle:
            return json.load(handle)

    def save_checkpoint(self, state: dict):
        self.paths.ensure_directories()
        state["updated_at"] = utc_now_iso()
        self.paths.checkpoint_path.write_text(
            json.dumps(state, indent=2, ensure_ascii=True),
            encoding="utf-8",
        )

    @staticmethod
    def _normalize_lean(value: str) -> str:
        normalized = (value or "").strip().lower().replace("_", "-")
        mapping = {
            "auth-left": "Auth-Left",
            "auth-right": "Auth-Right",
            "centrist": "Centrist",
            "lib-left": "Lib-Left",
            "lib-right": "Lib-Right",
        }
        return mapping.get(normalized, "")

    def build_batch_prompt(self, batch: list[dict]) -> str:
        numbered = "\n".join(f'{i}. "{row["evaluated_text"]}"' for i, row in enumerate(batch))
        return (
            f"{self.config.system_prompt}\n\n"
            "You will receive a numbered list of political statements.\n"
            "Return ONLY a JSON array in this exact shape:\n"
            "[\n"
            '  {"index": 0, "political_lean": "Auth-Left", "controversy_score_1_5": 3, '
            '"one_sentence_opinion": "One sentence only."}\n'
            "]\n\n"
            "Rules:\n"
            f"- political_lean must be exactly one of: {', '.join(ALLOWED_LEANS)}\n"
            "- controversy_score_1_5 must be an integer from 1 to 5\n"
            "- one_sentence_opinion must be exactly one sentence\n"
            "- Return one result for every input item\n"
            "- Do not include any prose outside the JSON array\n\n"
            f"Statements to evaluate:\n{numbered}"
        )

    def _error_result(self, record: dict, timestamp: str, error_message: str) -> dict:
        return {
            **record,
            "codex_political_lean": "",
            "controversy_score_1_5": "",
            "one_sentence_opinion": "",
            "model_name": self.config.model,
            "run_timestamp": timestamp,
            "status": "error",
            "error_message": error_message,
        }

    def evaluate_batch(self, batch: list[dict]) -> list[dict]:
        prompt = self.build_batch_prompt(batch)
        timestamp = utc_now_iso()
        try:
            payload = self.provider.call_json(prompt, schema_name="workbook-political-batch")
        except CodexProviderError as exc:
            logger.warning(
                "Batch call failed for records %s..%s: %s",
                batch[0]["record_id"],
                batch[-1]["record_id"],
                exc,
            )
            return [self._error_result(record, timestamp, str(exc)) for record in batch]

        if not isinstance(payload, list):
            return [self._error_result(record, timestamp, "Batch payload was not a JSON array.") for record in batch]

        by_index = {}
        for item in payload:
            if not isinstance(item, dict):
                continue
            index = item.get("index")
            if isinstance(index, int) and 0 <= index < len(batch):
                by_index[index] = item

        results = []
        for batch_index, record in enumerate(batch):
            item = by_index.get(batch_index)
            if item is None:
                results.append(self._error_result(record, timestamp, "Missing result for this statement."))
                continue
            lean = self._normalize_lean(str(item.get("political_lean", "")))
            score = item.get("controversy_score_1_5")
            opinion = str(item.get("one_sentence_opinion", "")).strip()
            if lean not in ALLOWED_LEANS:
                results.append(self._error_result(record, timestamp, "Invalid political_lean value."))
                continue
            if not isinstance(score, int) or not 1 <= score <= 5:
                results.append(self._error_result(record, timestamp, "Invalid controversy score."))
                continue
            if not opinion:
                results.append(self._error_result(record, timestamp, "Missing one_sentence_opinion value."))
                continue
            results.append(
                {
                    **record,
                    "codex_political_lean": lean,
                    "controversy_score_1_5": score,
                    "one_sentence_opinion": opinion,
                    "model_name": self.config.model,
                    "run_timestamp": timestamp,
                    "status": "success",
                    "error_message": "",
                }
            )
        return results

    def _pending_records(self, records: list[dict], results: dict) -> list[dict]:
        pending = []
        for record in records:
            cached = results.get(record["record_id"])
            if cached is None:
                pending.append(record)
                continue
            if self.config.retry_errors and cached.get("status") != "success":
                pending.append(record)
        return pending

    def _write_outputs(self, records: list[dict], results: dict, summary: dict | None = None):
        self.paths.ensure_directories()
        merged_rows = []
        for record in records:
            cached = results.get(record["record_id"])
            if cached:
                merged = {**record, **cached}
            else:
                merged = {
                    **record,
                    "codex_political_lean": "",
                    "controversy_score_1_5": "",
                    "one_sentence_opinion": "",
                    "model_name": self.config.model,
                    "run_timestamp": "",
                    "status": "pending",
                    "error_message": "",
                }
            merged_rows.append(merged)

        with self.paths.machine_output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=WORKBOOK_MACHINE_COLUMNS)
            writer.writeheader()
            for row in merged_rows:
                writer.writerow({column: row.get(column, "") for column in WORKBOOK_MACHINE_COLUMNS})

        excel_csv_path = self.paths.machine_output_path.with_name(
            f"{self.paths.machine_output_path.stem}_excel.csv"
        )
        with excel_csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=WORKBOOK_MACHINE_COLUMNS)
            writer.writeheader()
            for row in merged_rows:
                writer.writerow({column: row.get(column, "") for column in WORKBOOK_MACHINE_COLUMNS})

        with self.paths.review_output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=WORKBOOK_REVIEW_COLUMNS)
            writer.writeheader()
            for row in merged_rows:
                writer.writerow(
                    {
                        "record_id": row.get("record_id", ""),
                        "source_workbook": row.get("source_workbook", ""),
                        "language": row.get("language", ""),
                        "sheet_name": row.get("sheet_name", ""),
                        "variation_label": row.get("variation_label", ""),
                        "row_id": row.get("row_id", ""),
                        "original_text": row.get("original_text", ""),
                        "evaluated_text": row.get("evaluated_text", ""),
                        "category": row.get("category", ""),
                        "quadrant": row.get("quadrant", ""),
                        "codex_political_lean": row.get("codex_political_lean", ""),
                        "reference_match": str(row.get("quadrant", "") == row.get("codex_political_lean", "")),
                        "controversy_score_1_5": row.get("controversy_score_1_5", ""),
                        "one_sentence_opinion": row.get("one_sentence_opinion", ""),
                        "status": row.get("status", ""),
                        "error_message": row.get("error_message", ""),
                        "review_notes": "",
                    }
                )

        review_excel_csv_path = self.paths.review_output_path.with_name(
            f"{self.paths.review_output_path.stem}_excel.csv"
        )
        with review_excel_csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=WORKBOOK_REVIEW_COLUMNS)
            writer.writeheader()
            for row in merged_rows:
                writer.writerow(
                    {
                        "record_id": row.get("record_id", ""),
                        "source_workbook": row.get("source_workbook", ""),
                        "language": row.get("language", ""),
                        "sheet_name": row.get("sheet_name", ""),
                        "variation_label": row.get("variation_label", ""),
                        "row_id": row.get("row_id", ""),
                        "original_text": row.get("original_text", ""),
                        "evaluated_text": row.get("evaluated_text", ""),
                        "category": row.get("category", ""),
                        "quadrant": row.get("quadrant", ""),
                        "codex_political_lean": row.get("codex_political_lean", ""),
                        "reference_match": str(row.get("quadrant", "") == row.get("codex_political_lean", "")),
                        "controversy_score_1_5": row.get("controversy_score_1_5", ""),
                        "one_sentence_opinion": row.get("one_sentence_opinion", ""),
                        "status": row.get("status", ""),
                        "error_message": row.get("error_message", ""),
                        "review_notes": "",
                    }
                )

        if summary is not None:
            self.paths.summary_output_path.write_text(
                json.dumps(summary, indent=2, ensure_ascii=True),
                encoding="utf-8",
            )

    def run(self) -> dict:
        self.paths.ensure_directories()
        records = self.load_records()
        state = self.load_checkpoint()
        results = state.setdefault("results", {})
        pending = self._pending_records(records, results)
        logger.info("Loaded %s workbook records; %s pending.", len(records), len(pending))

        for start in range(0, len(pending), self.config.batch_size):
            batch = pending[start : start + self.config.batch_size]
            logger.info(
                "Evaluating workbook batch %s-%s of %s pending records.",
                start + 1,
                start + len(batch),
                len(pending),
            )
            batch_results = self.evaluate_batch(batch)
            for result in batch_results:
                results[result["record_id"]] = result
            self.save_checkpoint(state)

        summary = {"success": 0, "error": 0, "pending": 0, "total": len(records)}
        per_language = {}
        per_sheet = {}
        for record in records:
            result = results.get(record["record_id"])
            status = result.get("status", "pending") if result else "pending"
            summary[status] = summary.get(status, 0) + 1
            per_language.setdefault(record["language"], {"success": 0, "error": 0, "pending": 0, "total": 0})
            per_language[record["language"]]["total"] += 1
            per_language[record["language"]][status] += 1
            sheet_key = f'{record["language"]}::{record["sheet_name"]}'
            per_sheet.setdefault(sheet_key, {"success": 0, "error": 0, "pending": 0, "total": 0})
            per_sheet[sheet_key]["total"] += 1
            per_sheet[sheet_key][status] += 1
        summary["per_language"] = per_language
        summary["per_sheet"] = per_sheet
        self._write_outputs(records, results, summary=summary)
        return summary
