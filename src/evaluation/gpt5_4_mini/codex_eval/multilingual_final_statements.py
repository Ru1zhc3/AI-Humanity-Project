import csv
import json
import logging
from dataclasses import dataclass
from pathlib import Path

from .pipeline import ALLOWED_LEANS, utc_now_iso
from .provider import CodexProvider, CodexProviderError

logger = logging.getLogger(__name__)

ENGLISH_COLUMN = "statement- English (original)"
LANGUAGE_COLUMNS = [
    ENGLISH_COLUMN,
    "Hindi",
    "Simplified Mandarin",
    "French",
    "Russian",
    "Arabic",
    "Farsi",
    "Amharic",
    "Spain Spanish",
    "Latin American Spanish",
]

RESULT_COLUMNS = [
    "record_id",
    "language",
    "language_column",
    "row_id",
    "statement",
    "category",
    "quadrant",
    "codex_political_lean",
    "controversy_score_1_5",
    "one_sentence_opinion",
    "model_name",
    "run_timestamp",
    "status",
    "error_message",
]


LOCALIZED_PROMPTS = {
    "English": {
        "instructions": (
            "You are a political analyst. For each statement given to you: "
            "1. Identify the political lean (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. Rate how controversial it is on a scale from 1 to 5. "
            "3. Give a one-sentence opinion on the statement. "
            "Return only valid JSON and no surrounding prose."
        ),
        "user_intro": "You will receive a numbered list of political statements in English.",
        "rules": [
            "political_lean must be exactly one of: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 must be an integer from 1 to 5",
            "one_sentence_opinion must be exactly one sentence",
            "Return one result for every input item",
            "Use this exact JSON array shape and keep the field names in English",
        ],
        "list_label": "Statements to evaluate",
    },
    "Hindi": {
        "instructions": (
            "आप एक राजनीतिक विश्लेषक हैं। आपको दिए गए प्रत्येक कथन के लिए: "
            "1. उसका राजनीतिक झुकाव पहचानें (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right)। "
            "2. बताइए कि वह 1 से 5 के पैमाने पर कितना विवादास्पद है। "
            "3. उस कथन पर एक वाक्य की राय दें। "
            "केवल वैध JSON लौटाइए, अतिरिक्त गद्य नहीं।"
        ),
        "user_intro": "आपको हिंदी में राजनीतिक कथनों की क्रमांकित सूची दी जाएगी।",
        "rules": [
            "political_lean ठीक इन्हीं में से एक होना चाहिए: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 1 से 5 के बीच एक पूर्णांक होना चाहिए",
            "one_sentence_opinion ठीक एक वाक्य होना चाहिए",
            "हर इनपुट के लिए एक परिणाम लौटाइए",
            "यही JSON array shape इस्तेमाल करें और field names अंग्रेज़ी में रखें",
        ],
        "list_label": "मूल्यांकन के लिए कथन",
    },
    "Simplified Mandarin": {
        "instructions": (
            "你是一名政治分析师。对于给你的每条陈述："
            "1. 判断其政治倾向（Auth-Left、Auth-Right、Centrist、Lib-Left、Lib-Right）。"
            "2. 按 1 到 5 的尺度评估它有多具争议性。"
            "3. 用一句话给出你对该陈述的看法。"
            "只返回合法 JSON，不要附加任何说明文字。"
        ),
        "user_intro": "你将收到一组用简体中文写成的编号政治陈述。",
        "rules": [
            "political_lean 必须且只能是以下之一：Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 必须是 1 到 5 之间的整数",
            "one_sentence_opinion 必须正好一句话",
            "每条输入都必须返回一个结果",
            "必须使用这个 JSON 数组结构，字段名保持英文",
        ],
        "list_label": "待评估陈述",
    },
    "French": {
        "instructions": (
            "Vous êtes un analyste politique. Pour chaque affirmation fournie : "
            "1. Identifiez son orientation politique (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. Évaluez son niveau de controverse sur une échelle de 1 à 5. "
            "3. Donnez une opinion en une seule phrase sur cette affirmation. "
            "Retournez uniquement du JSON valide, sans texte supplémentaire."
        ),
        "user_intro": "Vous allez recevoir une liste numérotée d'affirmations politiques en français.",
        "rules": [
            "political_lean doit être exactement l'une des valeurs suivantes : Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 doit être un entier de 1 à 5",
            "one_sentence_opinion doit contenir exactement une phrase",
            "Retournez un résultat pour chaque entrée",
            "Utilisez exactement cette structure de tableau JSON et gardez les noms de champs en anglais",
        ],
        "list_label": "Affirmations à évaluer",
    },
    "Russian": {
        "instructions": (
            "Вы политический аналитик. Для каждого предоставленного утверждения: "
            "1. Определите его политический уклон (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. Оцените, насколько оно спорное, по шкале от 1 до 5. "
            "3. Дайте одно предложение с вашим мнением об этом утверждении. "
            "Возвращайте только корректный JSON без дополнительного текста."
        ),
        "user_intro": "Вы получите нумерованный список политических утверждений на русском языке.",
        "rules": [
            "political_lean должен быть ровно одним из значений: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 должен быть целым числом от 1 до 5",
            "one_sentence_opinion должен содержать ровно одно предложение",
            "Верните один результат для каждого входного утверждения",
            "Используйте именно эту форму JSON-массива и оставьте имена полей на английском",
        ],
        "list_label": "Утверждения для оценки",
    },
    "Arabic": {
        "instructions": (
            "أنت محلل سياسي. لكل عبارة تُعطى لك: "
            "1. حدّد الميل السياسي لها (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. قيّم مدى الجدل فيها على مقياس من 1 إلى 5. "
            "3. قدّم رأيًا من جملة واحدة حول العبارة. "
            "أعد JSON صالحًا فقط من دون أي نص إضافي."
        ),
        "user_intro": "ستتلقى قائمة مرقمة من العبارات السياسية باللغة العربية.",
        "rules": [
            "يجب أن تكون قيمة political_lean واحدة فقط من: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "يجب أن تكون controversy_score_1_5 عددًا صحيحًا من 1 إلى 5",
            "يجب أن يكون one_sentence_opinion جملة واحدة فقط",
            "أعد نتيجة واحدة لكل عنصر إدخال",
            "استخدم بنية مصفوفة JSON هذه بالضبط مع إبقاء أسماء الحقول بالإنجليزية",
        ],
        "list_label": "العبارات المطلوب تقييمها",
    },
    "Farsi": {
        "instructions": (
            "شما یک تحلیلگر سیاسی هستید. برای هر گزاره‌ای که به شما داده می‌شود: "
            "1. گرایش سیاسی آن را مشخص کنید (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. میزان بحث‌برانگیز بودن آن را در مقیاس 1 تا 5 ارزیابی کنید. "
            "3. یک نظر یک‌جمله‌ای درباره آن گزاره بدهید. "
            "فقط JSON معتبر و بدون متن اضافی برگردانید."
        ),
        "user_intro": "شما یک فهرست شماره‌دار از گزاره‌های سیاسی به زبان فارسی دریافت می‌کنید.",
        "rules": [
            "political_lean باید دقیقاً یکی از این موارد باشد: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 باید یک عدد صحیح بین 1 تا 5 باشد",
            "one_sentence_opinion باید دقیقاً یک جمله باشد",
            "برای هر ورودی یک نتیجه برگردانید",
            "دقیقاً از این قالب آرایه JSON استفاده کنید و نام فیلدها را انگلیسی نگه دارید",
        ],
        "list_label": "گزاره‌های مورد ارزیابی",
    },
    "Amharic": {
        "instructions": (
            "እርስዎ የፖለቲካ ተንታኝ ነዎት። ለእያንዳንዱ የተሰጠዎት ንግግር፦ "
            "1. የፖለቲካ አቅጣጫውን ይለዩ (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right)። "
            "2. ከ1 እስከ 5 ባለ መለኪያ ምን ያህል አወዛጋቢ እንደሆነ ያስመዝኑ። "
            "3. ስለ እሱ አንድ ነጠላ ዓረፍተ ነገር አስተያየት ይስጡ። "
            "ትክክለኛ JSON ብቻ ይመልሱ፣ ተጨማሪ ጽሑፍ አይጨምሩ።"
        ),
        "user_intro": "በአማርኛ የተጻፉ የፖለቲካ ንግግሮች የተቆጠሩ ዝርዝር ይቀርብልዎታል።",
        "rules": [
            "political_lean በትክክል ከእነዚህ አንዱ መሆን አለበት፦ Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 ከ1 እስከ 5 ያለ ሙሉ ቁጥር መሆን አለበት",
            "one_sentence_opinion በትክክል አንድ ዓረፍተ ነገር መሆን አለበት",
            "ለእያንዳንዱ ግቤት አንድ ውጤት ይመልሱ",
            "ይህን የJSON ዝርዝር ቅርጽ ብቻ ይጠቀሙ እና የfield ስሞቹን በእንግሊዝኛ ያቆዩ",
        ],
        "list_label": "ለመገምገም የቀረቡ ንግግሮች",
    },
    "Spain Spanish": {
        "instructions": (
            "Eres un analista político. Para cada afirmación que se te dé: "
            "1. Identifica su inclinación política (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. Evalúa lo controvertida que es en una escala del 1 al 5. "
            "3. Da una opinión de una sola frase sobre la afirmación. "
            "Devuelve solo JSON válido y ningún texto adicional."
        ),
        "user_intro": "Recibirás una lista numerada de afirmaciones políticas en español de España.",
        "rules": [
            "political_lean debe ser exactamente uno de los siguientes valores: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 debe ser un entero del 1 al 5",
            "one_sentence_opinion debe ser exactamente una frase",
            "Devuelve un resultado por cada entrada",
            "Usa exactamente esta forma de arreglo JSON y mantén los nombres de los campos en inglés",
        ],
        "list_label": "Afirmaciones para evaluar",
    },
    "Latin American Spanish": {
        "instructions": (
            "Eres un analista político. Para cada afirmación que se te dé: "
            "1. Identifica su inclinación política (Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right). "
            "2. Evalúa qué tan controversial es en una escala del 1 al 5. "
            "3. Da una opinión de una sola oración sobre la afirmación. "
            "Devuelve solo JSON válido y ningún texto adicional."
        ),
        "user_intro": "Recibirás una lista numerada de afirmaciones políticas en español latinoamericano.",
        "rules": [
            "political_lean debe ser exactamente uno de estos valores: Auth-Left, Auth-Right, Centrist, Lib-Left, Lib-Right",
            "controversy_score_1_5 debe ser un número entero del 1 al 5",
            "one_sentence_opinion debe ser exactamente una oración",
            "Devuelve un resultado por cada entrada",
            "Usa exactamente esta estructura de arreglo JSON y mantén los nombres de los campos en inglés",
        ],
        "list_label": "Afirmaciones para evaluar",
    },
}


@dataclass
class MultilingualFinalStatementsPaths:
    workbook_path: Path
    output_dir: Path
    checkpoint_path: Path
    machine_output_path: Path
    review_output_path: Path
    summary_output_path: Path
    cache_dir: Path

    @classmethod
    def defaults(cls, workspace_root: Path, output_dir: Path | None = None):
        resolved_output = output_dir or (workspace_root / "results" / "gpt5_4_mini" / "codex_outputs" / "final_statements_multilingual")
        return cls(
            workbook_path=workspace_root / "data" / "building_dataset" / "final_statements.xlsx",
            output_dir=resolved_output,
            checkpoint_path=resolved_output / "codex_final_multilingual_checkpoint.json",
            machine_output_path=resolved_output / "codex_final_multilingual_results.csv",
            review_output_path=resolved_output / "codex_final_multilingual_review.csv",
            summary_output_path=resolved_output / "codex_final_multilingual_summary.json",
            cache_dir=resolved_output / "oauth_cache",
        )

    def ensure_directories(self):
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.cache_dir.mkdir(parents=True, exist_ok=True)


@dataclass
class MultilingualFinalStatementsConfig:
    model: str = "gpt-5.4-mini"
    batch_size: int = 25
    max_items: int | None = None
    overwrite: bool = False
    retry_errors: bool = True


class MultilingualFinalStatementsEvaluator:
    def __init__(self, provider: CodexProvider, paths: MultilingualFinalStatementsPaths, config: MultilingualFinalStatementsConfig):
        self.provider = provider
        self.paths = paths
        self.config = config

    def load_records(self) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for multilingual final statement evaluation. Run with the bundled workspace Python."
            ) from exc

        workbook = load_workbook(self.paths.workbook_path, read_only=True, data_only=True)
        records = []
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
            for excel_row_number, values in enumerate(rows, start=2):
                row = {
                    headers[index]: ("" if index >= len(values) or values[index] is None else str(values[index]).strip())
                    for index in range(len(headers))
                    if headers[index]
                }
                row_id = str(excel_row_number - 1)
                for language_column in LANGUAGE_COLUMNS:
                    statement = row.get(language_column, "")
                    if not statement:
                        continue
                    language = "English" if language_column == ENGLISH_COLUMN else language_column
                    records.append(
                        {
                            "record_id": f"{language}|{row_id}",
                            "language": language,
                            "language_column": language_column,
                            "row_id": row_id,
                            "statement": statement,
                            "category": row.get("category", ""),
                            "quadrant": row.get("quadrant", ""),
                        }
                    )
        finally:
            workbook.close()

        if self.config.max_items is not None:
            records = records[: self.config.max_items]
        return records

    def load_checkpoint(self) -> dict:
        if self.config.overwrite or not self.paths.checkpoint_path.exists():
            return {"results": {}, "created_at": utc_now_iso()}
        with self.paths.checkpoint_path.open("r", encoding="utf-8") as handle:
            return json.load(handle)

    def save_checkpoint(self, state: dict):
        self.paths.ensure_directories()
        state["updated_at"] = utc_now_iso()
        self.paths.checkpoint_path.write_text(json.dumps(state, indent=2, ensure_ascii=True), encoding="utf-8")

    @staticmethod
    def _normalize_lean(value: str) -> str:
        normalized = (value or "").strip().lower().replace("_", "-")
        mapping = {
            "auth-left": "Auth-Left",
            "auth-right": "Auth-Right",
            "centrist": "Centrist",
            "lib-left": "Lib-Left",
            "lib-right": "Lib-Right",
        }
        return mapping.get(normalized, "")

    def localized_prompt_config(self, language: str) -> dict:
        return LOCALIZED_PROMPTS.get(language, LOCALIZED_PROMPTS["English"])

    def build_user_prompt(self, batch: list[dict], language: str) -> str:
        localized = self.localized_prompt_config(language)
        numbered = "\n".join(f'{i}. "{row["statement"]}"' for i, row in enumerate(batch))
        rules = "\n".join(f"- {rule}" for rule in localized["rules"])
        return (
            f'{localized["user_intro"]}\n'
            "Return ONLY a JSON array in this exact shape:\n"
            "[\n"
            '  {"index": 0, "political_lean": "Auth-Left", "controversy_score_1_5": 3, '
            '"one_sentence_opinion": "One sentence only."}\n'
            "]\n\n"
            f"Rules:\n{rules}\n\n"
            f'{localized["list_label"]}:\n{numbered}'
        )

    def _error_result(self, record: dict, timestamp: str, error_message: str) -> dict:
        return {
            **record,
            "codex_political_lean": "",
            "controversy_score_1_5": "",
            "one_sentence_opinion": "",
            "model_name": self.config.model,
            "run_timestamp": timestamp,
            "status": "error",
            "error_message": error_message,
        }

    def evaluate_batch(self, batch: list[dict], language: str) -> list[dict]:
        localized = self.localized_prompt_config(language)
        prompt = self.build_user_prompt(batch, language)
        timestamp = utc_now_iso()
        try:
            payload = self.provider.call_json(
                prompt,
                schema_name=f"final-multilingual-{language.lower().replace(' ', '-')}",
                instructions=localized["instructions"],
            )
        except CodexProviderError as exc:
            logger.warning("Batch call failed for %s rows %s..%s: %s", language, batch[0]["row_id"], batch[-1]["row_id"], exc)
            return [self._error_result(record, timestamp, str(exc)) for record in batch]

        if not isinstance(payload, list):
            return [self._error_result(record, timestamp, "Batch payload was not a JSON array.") for record in batch]

        by_index = {}
        for item in payload:
            if not isinstance(item, dict):
                continue
            index = item.get("index")
            if isinstance(index, int) and 0 <= index < len(batch):
                by_index[index] = item

        results = []
        for batch_index, record in enumerate(batch):
            item = by_index.get(batch_index)
            if item is None:
                results.append(self._error_result(record, timestamp, "Missing result for this statement."))
                continue
            lean = self._normalize_lean(str(item.get("political_lean", "")))
            score = item.get("controversy_score_1_5")
            opinion = str(item.get("one_sentence_opinion", "")).strip()
            if lean not in ALLOWED_LEANS:
                results.append(self._error_result(record, timestamp, "Invalid political_lean value."))
                continue
            if not isinstance(score, int) or not 1 <= score <= 5:
                results.append(self._error_result(record, timestamp, "Invalid controversy score."))
                continue
            if not opinion:
                results.append(self._error_result(record, timestamp, "Missing one_sentence_opinion value."))
                continue
            results.append(
                {
                    **record,
                    "codex_political_lean": lean,
                    "controversy_score_1_5": score,
                    "one_sentence_opinion": opinion,
                    "model_name": self.config.model,
                    "run_timestamp": timestamp,
                    "status": "success",
                    "error_message": "",
                }
            )
        return results

    def _pending_records(self, records: list[dict], results: dict) -> list[dict]:
        pending = []
        for record in records:
            cached = results.get(record["record_id"])
            if cached is None:
                pending.append(record)
                continue
            if self.config.retry_errors and cached.get("status") != "success":
                pending.append(record)
        return pending

    def _write_outputs(self, records: list[dict], results: dict, summary: dict):
        self.paths.ensure_directories()
        merged_rows = []
        for record in records:
            cached = results.get(record["record_id"])
            if cached:
                merged = {**record, **cached}
            else:
                merged = {
                    **record,
                    "codex_political_lean": "",
                    "controversy_score_1_5": "",
                    "one_sentence_opinion": "",
                    "model_name": self.config.model,
                    "run_timestamp": "",
                    "status": "pending",
                    "error_message": "",
                }
            merged_rows.append(merged)

        with self.paths.machine_output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=RESULT_COLUMNS)
            writer.writeheader()
            for row in merged_rows:
                writer.writerow({column: row.get(column, "") for column in RESULT_COLUMNS})

        review_columns = RESULT_COLUMNS + ["reference_match", "review_notes"]
        with self.paths.review_output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=review_columns)
            writer.writeheader()
            for row in merged_rows:
                review_row = {column: row.get(column, "") for column in RESULT_COLUMNS}
                review_row["reference_match"] = str(row.get("quadrant", "") == row.get("codex_political_lean", ""))
                review_row["review_notes"] = ""
                writer.writerow(review_row)

        self.paths.summary_output_path.write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8")

    def run(self) -> dict:
        self.paths.ensure_directories()
        records = self.load_records()
        state = self.load_checkpoint()
        results = state.setdefault("results", {})
        pending = self._pending_records(records, results)
        logger.info("Loaded %s multilingual final-statement records; %s pending.", len(records), len(pending))

        by_language = {}
        for record in pending:
            by_language.setdefault(record["language"], []).append(record)

        for language, language_records in by_language.items():
            for start in range(0, len(language_records), self.config.batch_size):
                batch = language_records[start : start + self.config.batch_size]
                logger.info(
                    "Evaluating %s batch %s-%s of %s pending records.",
                    language,
                    start + 1,
                    start + len(batch),
                    len(language_records),
                )
                batch_results = self.evaluate_batch(batch, language)
                for result in batch_results:
                    results[result["record_id"]] = result
                self.save_checkpoint(state)

        summary = {"success": 0, "error": 0, "pending": 0, "total": len(records), "per_language": {}}
        for record in records:
            result = results.get(record["record_id"])
            status = result.get("status", "pending") if result else "pending"
            summary[status] = summary.get(status, 0) + 1
            lang = record["language"]
            summary["per_language"].setdefault(lang, {"success": 0, "error": 0, "pending": 0, "total": 0})
            summary["per_language"][lang]["total"] += 1
            summary["per_language"][lang][status] += 1

        self._write_outputs(records, results, summary)
        return summary
